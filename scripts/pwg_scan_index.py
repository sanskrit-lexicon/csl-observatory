#!/usr/bin/env python3
"""PWG literary-source scan-index campaign — registry, coverage, and e-text queue (H1706).

Between January 2025 and July 2026 a volunteer team page-indexed the printed
literary sources cited in the Boehtlingk-Roth Sanskrit-Woerterbuch (PWG), so that a
`<ls>` citation in the dictionary can be resolved to a page image of the edition it
cites. The campaign was tracked in a Google Sheet; this script turns that sheet into
committed, machine-readable data and measures the campaign the way the dictionary
cares about it -- by **citation mass**, not by row count.

Three outputs, three questions:

1. `data/pwg_scan_index_tracker/pwg_scan_index.tsv` -- the registry. One row per
   tracked literary source: LS code, citation count, pages, status, volunteer, dates,
   coordinating issue, scan directory. Plus a validation column set (see below).
2. `reports/pwg_scan_index.md` -- the coverage surface. What share of PWG's citation
   apparatus is page-indexed, what remains, per-volunteer throughput, velocity.
3. `data/pwg_scan_index_tracker/pwg_etext_candidate_queue.tsv` -- the ranked queue of
   done-and-indexed works that are candidates for full-text extraction (H1689 recipe).

Cross-validation (divergences are recorded as FINDINGS, never written back as
corrections -- the sheet is a human work-log, the dictionary data is a separate
measurement, and where they disagree that disagreement is the interesting part):

* LS entry codes vs the PWG abbreviation bibliography `pwgbib_input.txt`
  (sanskrit-lexicon/PWG `pwg_ls1/pwgauth/`).
* Citation counts vs `sortedcrefs.txt` (sanskrit-lexicon/PWG `pwg_ls/pwg_dhaval/`),
  the full-dictionary `<ls>` extraction -- the tracker's own denominator.
* Citation counts vs `citation_sources.json` (sanskrit-lexicon/PWG
  `pwg_ls/pwg_ru_coverage/`), which counts only the translated article subset and so
  is expected to be much smaller; it is reported for scale, not as a check.

Inputs are the committed CSV snapshot under `data/pwg_scan_index_tracker/snapshot/`.
`--fetch` re-downloads the live sheet (all tabs) and rewrites that snapshot first;
without it the script is stdlib-only and fully offline.

Run:    python scripts/pwg_scan_index.py
        python scripts/pwg_scan_index.py --fetch      # refresh the snapshot first
        python scripts/pwg_scan_index.py --pwg-repo <path to a PWG checkout>
"""

import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

SHEET_ID = "1rcYfQE0D26RNdWSmRQzhFnV3Gf248wSuldTj-wt8_O0"
SHEET_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/edit?gid=0"
XLSX_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

DEFAULT_PWG_REPO = r"C:\Users\user\Documents\GitHub\PWG"

REPO_BLOB = "https://github.com/sanskrit-lexicon/csl-observatory/blob/main"
PWG_BLOB = "https://github.com/sanskrit-lexicon/PWG/blob/main"
SCANS_ORG = "https://github.com/sanskrit-lexicon-scans"

# Placeholders the sheet uses for "no value here" -- three spaced hyphens for a value
# that belongs to the parent row, four spaced dots for "not applicable to this row".
NULLISH = {"", "- - -", ". . . .", "-", "--", "n/a"}

# Index Status vocabulary as written in the sheet -> normalized slug + gloss.
STATUS_MAP = {
    "done": ("done", "index built, reviewed, and posted"),
    "on-going": ("on-going", "reserved and in progress"),
    "to do/ open": ("to-do", "unclaimed, no volunteer assigned"),
    "to do/open": ("to-do", "unclaimed, no volunteer assigned"),
    "page-wise": ("page-wise", "cited page-wise; no per-entry index needed"),
    "nr/ indirect": ("nr-indirect", "not required -- cited only indirectly"),
    "nr/ alt. name": ("nr-alt-name", "not required -- an alternate abbreviation of another row"),
}
DONE_STATUSES = {"done"}


# --------------------------------------------------------------------------- fetch

# Columns redacted before the snapshot is committed, by tab slug -> column indices.
# The sheet's `Team` tab maps each volunteer's real personal name to their GitHub
# handle. The handles are public -- they appear on every coordinating issue -- but
# the name-to-handle linkage is not, and this repository is public. Committing it
# would publish personal data beyond what the public record already carries, which
# is not a call an automated pass gets to make. The handle column is kept, so the
# roster and every credit in the reports survive the redaction intact.
REDACT_COLUMNS = {"team": [2]}
REDACTED = "[redacted: personal name]"


def fetch_snapshot(snapshot_dir: Path) -> None:
    """Download every tab of the live sheet as CSV into snapshot_dir."""
    import urllib.request

    from openpyxl import load_workbook  # only needed on the --fetch path

    snapshot_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = snapshot_dir / "_source.xlsx"
    print(f"fetching {XLSX_URL}")
    with urllib.request.urlopen(XLSX_URL, timeout=120) as resp:
        xlsx_path.write_bytes(resp.read())

    wb = load_workbook(xlsx_path, data_only=True)
    for name in wb.sheetnames:
        ws = wb[name]
        slug = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")
        out = snapshot_dir / f"{slug}.csv"
        redact = REDACT_COLUMNS.get(slug, [])
        n_redacted = 0
        with out.open("w", encoding="utf-8", newline="") as fh:
            w = csv.writer(fh)
            for row in ws.iter_rows(values_only=True):
                cells = ["" if c is None else str(c) for c in row]
                for i in redact:
                    if i < len(cells) and cells[i].strip():
                        cells[i] = REDACTED
                        n_redacted += 1
                w.writerow(cells)
        note = f", {n_redacted} cells redacted" if n_redacted else ""
        print(f"  wrote {out.name} ({ws.max_row} rows{note})")
    xlsx_path.unlink()


# ----------------------------------------------------------------------- normalize

def clean(value) -> str:
    s = "" if value is None else str(value).strip()
    return "" if s.lower() in NULLISH else s


def as_int(value):
    s = clean(value)
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def as_date(value):
    s = clean(value)
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def iso(d):
    return d.isoformat() if d else ""


def normalize_code(raw: str):
    """Strip the sheet's presentation marks off an LS entry code.

    Returns (code, vedic_marked, volume_part). The leading black star marks the
    Vedic group; ' vol.N' and ' (N vol.s)' are physical-volume bookkeeping, not part
    of the abbreviation as PWG prints it.
    """
    code = raw.strip()
    vedic = code.startswith("\u2605")
    code = code.lstrip("\u2605").strip()
    volume = ""
    m = re.search(r"\s+vol\.\s*(\d+)$", code)
    if m:
        volume = m.group(1)
        code = code[: m.start()].strip()
    m = re.search(r"\s*\((\d+)\s*vol\.s\)$", code)
    if m:
        volume = volume or f"{m.group(1)} vols"
        code = code[: m.start()].strip()
    return code, vedic, volume


def base_code(code: str) -> str:
    """The bare abbreviation as pwgbib would list it (drop edition parentheticals)."""
    out = re.sub(r"\s*\([^)]*\)\s*$", "", code).strip()
    return out or code


def unqualify(code: str) -> str:
    """Strip every edition/volume qualification the sheet adds to an abbreviation.

    `Spr. vol.1 (1st ed.)` -> `Spr.`, `AK. Deslongchamps ed.` -> `AK.`,
    `MBh. (Bomb.)` -> `MBh.`, `TS. (2 vol.s)` -> `TS.`
    """
    out = base_code(code)
    out = re.sub(r"\s+vol\.\s*\d+$", "", out).strip()
    out = re.sub(r"\s+\S+(\s+\S+)?\s+ed\.$", "", out).strip()
    return out or code


def match_pwgbib(code: str, pwgbib: dict, pwgbib_ci: dict):
    """Resolve a sheet LS code against the bibliography, cheapest test first.

    Returns (matched key, tier). The tiers matter: `exact` means the sheet and the
    bibliography spell the abbreviation identically; `case` and `qualified` mean the
    work is unambiguously known but written differently; `none` is a real loose end.
    """
    if code in pwgbib:
        return code, "exact"
    if code.lower() in pwgbib_ci:
        return pwgbib_ci[code.lower()], "case"
    bare = unqualify(code)
    if bare != code:
        if bare in pwgbib:
            return bare, "qualified"
        if bare.lower() in pwgbib_ci:
            return pwgbib_ci[bare.lower()], "qualified-case"
    return "", "none"


def parse_works(path: Path):
    rows = []
    with path.open(encoding="utf-8", newline="") as fh:
        raw = list(csv.reader(fh))
    header = raw[0]
    assert "Index Status" in [c.strip() for c in header], f"unexpected header: {header}"
    for r in raw[1:]:
        if not any(clean(c) for c in r):
            continue
        r = r + [""] * (15 - len(r))
        book_no_raw = clean(r[0])
        # '*10' marks a work sharing a physical book with the row above it.
        shares_book = book_no_raw.startswith("*")
        book_no = book_no_raw.lstrip("*")
        if re.fullmatch(r"\d+\.0", book_no):
            book_no = book_no[:-2]
        code_raw = clean(r[2])
        if not code_raw:
            continue
        code, vedic, volume = normalize_code(code_raw)
        status_raw = clean(r[5]).lower()
        status, status_gloss = STATUS_MAP.get(status_raw, (status_raw or "unknown", ""))
        issue_raw = clean(r[9])
        issue_repo, issue_no = "", None
        m = re.match(r"(PWG|PWK)/issues/(\d+)", issue_raw)
        if m:
            issue_repo, issue_no = m.group(1), int(m.group(2))
        scan_dir = clean(r[11]).lstrip("/")
        pages_raw = clean(r[3])
        rows.append({
            "book_no": book_no,
            "shares_book_with_previous": shares_book,
            "ls_code": code,
            "ls_code_raw": code_raw,
            "ls_code_base": base_code(code),
            "vedic_marked": vedic,
            "volume": volume,
            "title": clean(r[4]),
            "citation_count": as_int(r[1]),
            "total_pages": as_int(pages_raw),
            "total_pages_raw": pages_raw,
            "status": status,
            "status_gloss": status_gloss,
            "status_raw": clean(r[5]),
            "volunteer": clean(r[6]),
            "started": as_date(r[7]),
            "finished": as_date(r[8]),
            "issue_repo": issue_repo,
            "issue_no": issue_no,
            "index_posted": as_date(r[10]),
            "scan_dir": scan_dir,
            "scan_dir_raw": clean(r[11]),
            "public_link": as_date(r[12]),
            "sheet_row_no": as_int(r[13]),
        })
    return rows


def parse_simple_csv(path: Path):
    if not path.exists():
        return []
    with path.open(encoding="utf-8", newline="") as fh:
        return list(csv.reader(fh))


# ------------------------------------------------------------------- cross-checks

def load_pwgbib(pwg_repo: Path):
    """{code: description} from the PWG abbreviation bibliography."""
    path = pwg_repo / "pwg_ls1" / "pwgauth" / "pwgbib_input.txt"
    if not path.exists():
        return {}, path
    out = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        parts = line.split("\t")
        if len(parts) >= 4:
            out.setdefault(parts[1].strip(), parts[3].strip())
    return out, path


def load_sortedcrefs(pwg_repo: Path):
    """{citation-string: occurrences} over the WHOLE dictionary."""
    path = pwg_repo / "pwg_ls" / "pwg_dhaval" / "abbrvwork" / "abbrvoutput" / "sortedcrefs.txt"
    if not path.exists():
        return {}, 0, path
    out = {}
    total = 0
    for line in path.read_text(encoding="utf-8").splitlines():
        parts = line.rsplit("@", 2)
        if len(parts) != 3:
            continue
        try:
            n = int(parts[2])
        except ValueError:
            continue
        out[parts[0].strip()] = out.get(parts[0].strip(), 0) + n
        total += n
    return out, total, path


def load_ru_subset_counts(pwg_repo: Path):
    """Resolver evidence from the translated-article-subset coverage measurement.

    Returns ({abbr: {...}}, {scan-repo names the resolver actually emitted}, path).
    The `repos` field of each entry names the Cologne scan/HTML repository the
    resolver pointed that abbreviation at, so the union of those names is a
    *observed* list of wired targets — observed over the translated subset only,
    which is why absence from it is evidence of an unwired index, not proof.
    """
    path = pwg_repo / "pwg_ls" / "pwg_ru_coverage" / "citation_sources.json"
    if not path.exists():
        return {}, set(), path
    data = json.loads(path.read_text(encoding="utf-8"))
    out, repos_seen = {}, set()
    for item in data.get("abbreviations") or []:
        if not isinstance(item, dict):
            continue
        key = item.get("abbr") or item.get("abbrev") or item.get("abbreviation") or item.get("code")
        if not key:
            continue
        repos = [r for r in (item.get("repos") or []) if r]
        out[key] = {
            "total": item.get("total") or item.get("count") or 0,
            "scan": item.get("scan") or 0,
            "html": item.get("html") or 0,
            "repos": repos,
            "sample_url": item.get("sample_url") or "",
        }
        repos_seen.update(repos)
    return out, repos_seen, path


def load_scan_audit(tracker_dir: Path):
    """The dated audit of every scan directory: does the repo exist, is the resolver
    wired to it, and does the tracker spell it the way GitHub Pages needs.

    Hand-built once (see this directory's README for provenance) rather than
    recomputed per run, because it costs live GitHub API calls and a static read of
    a resolver that lives in another repository. Keyed case-insensitively -- the
    tracker's own casing is one of the defects it records.
    """
    path = tracker_dir / "scan_target_audit.tsv"
    if not path.exists():
        return {}, path
    out = {}
    with path.open(encoding="utf-8", newline="") as fh:
        for row in csv.DictReader(fh, delimiter="\t"):
            out[row["scan_dir"].lower()] = row
            out[row["tracker_spelling"].lower()] = row
    return out, path


def load_provenance():
    """Per-row citation-count provenance, from the H2874 dataset.

    Built by `scripts/pwg_citation_count_provenance.py`, which also owns the contract
    gate. Absent file -> empty, and the columns below simply stay blank; this generator
    must not become the thing that decides what a tracker number means.
    """
    path = (Path(__file__).resolve().parents[1] / "data" / "pwg_scan_index_tracker"
            / "pwg_citation_count_provenance.tsv")
    if not path.exists():
        print(f"WARNING: no provenance dataset at {path}; citation-count provenance "
              "columns will be blank", file=sys.stderr)
        return []
    with path.open(encoding="utf-8") as fh:
        return list(csv.DictReader(fh, delimiter="\t"))


def load_ls_count_denominators():
    """(tracker-era ALL, current ALL) from the committed ls-count table metadata."""
    base = (Path(__file__).resolve().parents[1] / "data" / "pwg_scan_index_tracker"
            / "ls_counts")
    out = []
    for name in ("pwg_ls_counts_2024-09-11.meta.json", "pwg_ls_counts_current.meta.json"):
        p = base / name
        out.append(json.loads(p.read_text(encoding="utf-8"))["totals"]["ALL"]
                   if p.exists() else None)
    return tuple(out)


def lookup_variants(row):
    """Code spellings to try against an external table, most specific first."""
    seen, out = set(), []
    for cand in (row["ls_code"], row["ls_code_base"], row["ls_code_raw"].lstrip("\u2605").strip()):
        cand = cand.strip()
        if cand and cand not in seen:
            seen.add(cand)
            out.append(cand)
    return out


# ------------------------------------------------------------------------- report

def pct(n, d):
    return round(100.0 * n / d, 1) if d else 0.0


def fmt(n):
    return f"{n:,}" if isinstance(n, int) else ""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--fetch", action="store_true",
                    help="re-download the live sheet into the committed snapshot first")
    ap.add_argument("--pwg-repo", default=DEFAULT_PWG_REPO,
                    help="path to a sanskrit-lexicon/PWG checkout for cross-validation")
    ap.add_argument("--outdir", default=str(Path(__file__).resolve().parent.parent))
    args = ap.parse_args()

    outdir = Path(args.outdir)
    tracker_dir = outdir / "data" / "pwg_scan_index_tracker"
    snapshot_dir = tracker_dir / "snapshot"
    reports_dir = outdir / "reports"
    site_data = outdir / "observatory" / "site" / "src" / "data"
    for d in (tracker_dir, reports_dir, site_data):
        d.mkdir(parents=True, exist_ok=True)

    if args.fetch:
        fetch_snapshot(snapshot_dir)

    works_csv = snapshot_dir / "works.csv"
    if not works_csv.exists():
        print(f"BLOCKED: no snapshot at {works_csv}; run with --fetch", file=sys.stderr)
        sys.exit(1)

    rows = parse_works(works_csv)
    pwg_repo = Path(args.pwg_repo)
    pwgbib, pwgbib_path = load_pwgbib(pwg_repo)
    crefs, crefs_total, crefs_path = load_sortedcrefs(pwg_repo)
    ru_counts, resolver_repos, ru_path = load_ru_subset_counts(pwg_repo)
    scan_audit, audit_path = load_scan_audit(tracker_dir)

    # ---- annotate each row with the cross-checks -------------------------------
    pwgbib_ci = {}
    for k in pwgbib:
        pwgbib_ci.setdefault(k.lower(), k)
    for r in rows:
        key, tier = match_pwgbib(r["ls_code"], pwgbib, pwgbib_ci)
        r["in_pwgbib"] = key
        r["pwgbib_tier"] = tier
        r["pwgbib_desc"] = pwgbib.get(key, "")
        # Occurrences of this abbreviation as a BARE cleaned citation string in the
        # full-dictionary extraction. Only an exact hit on the code as the sheet
        # writes it counts: falling back to a shorter parent form would silently
        # credit `R. (Bomb.)` with every `R.` citation in the dictionary, which is a
        # different recension. Rows whose code carries an edition qualifier are
        # therefore left unmatched on purpose.
        r["dict_citations"] = None
        r["dict_citations_key"] = ""
        for cand in (r["ls_code"], r["ls_code_raw"].lstrip("★").strip()):
            if cand in crefs:
                r["dict_citations"] = crefs[cand]
                r["dict_citations_key"] = cand
                break
        r["edition_qualified"] = bool(re.search(r"\([^)]*\)", r["ls_code"]))
        r["ru_subset_refs"] = None
        r["ru_subset_scan"] = None
        r["resolver_repos"] = ""
        r["ru_match_key"] = ""
        for cand in lookup_variants(r):
            if cand in ru_counts:
                hit = ru_counts[cand]
                r["ru_subset_refs"] = hit["total"]
                r["ru_subset_scan"] = hit["scan"]
                r["resolver_repos"] = ";".join(hit["repos"])
                # Record WHICH spelling matched. Where it is not the row's own code,
                # the count belongs to the parent abbreviation and is shared by every
                # edition/volume row under it -- do not add such rows together.
                r["ru_match_key"] = cand
                break
        # Wiring verdict. The dated static audit is authoritative where it exists --
        # it read the resolver source directly. The translated-subset observation is
        # only a fallback, and a weak one: absence there can mean "never cited in
        # that subset" rather than "no target".
        aud = scan_audit.get(r["scan_dir"].lower()) if r["scan_dir"] else None
        r["scan_dir_canonical"] = aud["scan_dir"] if aud else r["scan_dir"]
        r["scan_spelling_ok"] = aud["spelling_matches"] if aud else ""
        r["scan_repo_exists"] = aud["repo_exists"] if aud else ""
        r["scan_pages_url"] = aud["pages_url"] if aud else ""
        r["scan_repo_size_kb"] = aud["size_kb"] if aud else ""
        r["scan_repo_pushed_at"] = aud["pushed_at"] if aud else ""
        r["scan_audit_note"] = aud["note"] if aud else ""
        if aud:
            r["scan_wired"] = aud["resolver_wired"]
        elif r["scan_dir"]:
            r["scan_wired"] = "observed" if r["scan_dir"] in resolver_repos else "unaudited"
        else:
            r["scan_wired"] = ""
        # Ratio, not verdict. The two numbers count different objects (a book vs a
        # cleaned citation string), so "agrees / disagrees" is the wrong question --
        # see the report's cross-validation section.
        sheet_n, dict_n = r["citation_count"], r["dict_citations"]
        if sheet_n and dict_n:
            r["count_ratio"] = round(sheet_n / dict_n, 2)
        else:
            r["count_ratio"] = None

    # ---- citation-count provenance (H2874) -------------------------------------
    # The sheet's column is the 2024-09-11 `lsextract_all.txt` rollup; `citation_count`
    # keeps the sheet's own transcription, `citation_count_safe` carries the number that
    # table actually holds, and only the safe field is allowed into arithmetic.
    prov = {p["ls_code"]: p for p in load_provenance()}
    ls_all_tracker_era, ls_all_current = load_ls_count_denominators()
    for r in rows:
        p = prov.get(r["ls_code"], {})
        r["citation_count_provenance"] = p.get("provenance", "")
        r["citation_count_safe"] = int(p["citation_count_safe"]) \
            if p.get("citation_count_safe") else None
        r["citation_count_full"] = int(p["canonical_value"]) \
            if p.get("canonical_value") else None

    # ---- aggregate -------------------------------------------------------------
    by_status = defaultdict(list)
    for r in rows:
        by_status[r["status"]].append(r)

    # Citation mass runs on the provenance-checked value, not the sheet's transcription
    # of it (H2874). The two differ on one row today; the point is that they are allowed
    # to differ at all, and only one of them has a denominator behind it.
    def mass(rs):
        if prov:
            return sum(r["citation_count_safe"] or 0 for r in rs)
        return sum(r["citation_count"] or 0 for r in rs)

    def pages(rs):
        return sum(r["total_pages"] or 0 for r in rs)

    tracked_mass = mass(rows)
    done_rows = by_status["done"]
    indexed_mass = mass(done_rows)

    # Per-volunteer throughput. A volunteer cell can carry a parenthetical second
    # handle ("@IrinaKonstant (@195629012025)"); the first handle is the person.
    def first_handle(s):
        m = re.match(r"(@[\w.-]+)", s.strip())
        return m.group(1) if m else s.strip()

    per_vol = defaultdict(lambda: {"rows": 0, "mass": 0, "pages": 0, "works": []})
    for r in rows:
        if not r["volunteer"]:
            continue
        h = first_handle(r["volunteer"])
        per_vol[h]["rows"] += 1
        per_vol[h]["mass"] += mass([r])
        per_vol[h]["pages"] += r["total_pages"] or 0
        per_vol[h]["works"].append(r["ls_code"])

    # Velocity: indexes finished per month, public links per month.
    fin_month = Counter(r["finished"].strftime("%Y-%m") for r in rows if r["finished"])
    pub_month = Counter(r["public_link"].strftime("%Y-%m") for r in rows if r["public_link"])
    months = sorted(set(fin_month) | set(pub_month))

    # Turnaround: index posted -> public scan link live.
    lags = [(r["public_link"] - r["index_posted"]).days
            for r in rows if r["public_link"] and r["index_posted"]]
    lags.sort()
    lag_median = lags[len(lags) // 2] if lags else None

    todo = sorted([r for r in rows if r["status"] == "to-do"],
                  key=lambda r: -(r["citation_count"] or 0))
    ongoing = sorted([r for r in rows if r["status"] == "on-going"],
                     key=lambda r: -(r["citation_count"] or 0))
    pagewise = sorted([r for r in rows if r["status"] == "page-wise"],
                      key=lambda r: -(r["citation_count"] or 0))
    nr = [r for r in rows if r["status"].startswith("nr")]

    # ---- registry TSV ----------------------------------------------------------
    cols = ["ls_code", "ls_code_raw", "ls_code_base", "book_no", "volume", "title",
            "status", "status_gloss", "vedic_marked", "shares_book_with_previous",
            "citation_count", "citation_count_safe", "citation_count_provenance",
            "citation_count_full", "total_pages", "volunteer", "started", "finished",
            "issue_repo", "issue_no", "issue_url", "index_posted", "scan_dir",
            "scan_dir_canonical", "scan_spelling_ok", "scan_url", "scan_pages_url",
            "scan_repo_exists", "scan_repo_size_kb", "scan_repo_pushed_at",
            "scan_audit_note", "public_link", "in_pwgbib", "dict_citations",
            "dict_citations_key", "count_ratio", "edition_qualified", "ru_subset_refs",
            "ru_subset_scan", "ru_match_key", "resolver_repos", "scan_wired",
            "sheet_row_no"]
    reg_path = tracker_dir / "pwg_scan_index.tsv"
    with reg_path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, delimiter="\t", lineterminator="\n")
        w.writerow(cols)
        for r in rows:
            issue_url = (f"https://github.com/sanskrit-lexicon/{r['issue_repo']}/issues/{r['issue_no']}"
                         if r["issue_no"] else "")
            # Always build the URL from the CANONICAL directory name: GitHub Pages
            # paths are case-sensitive, and the tracker's own casing is wrong for one.
            scan_url = f"{SCANS_ORG}/{r['scan_dir_canonical']}" if r["scan_dir_canonical"] else ""
            w.writerow([
                r["ls_code"], r["ls_code_raw"], r["ls_code_base"], r["book_no"], r["volume"],
                r["title"], r["status"], r["status_gloss"],
                "yes" if r["vedic_marked"] else "", "yes" if r["shares_book_with_previous"] else "",
                r["citation_count"] if r["citation_count"] is not None else "",
                r["citation_count_safe"] if r["citation_count_safe"] is not None else "",
                r["citation_count_provenance"],
                r["citation_count_full"] if r["citation_count_full"] is not None else "",
                r["total_pages"] if r["total_pages"] is not None else "",
                r["volunteer"], iso(r["started"]), iso(r["finished"]),
                r["issue_repo"], r["issue_no"] or "", issue_url, iso(r["index_posted"]),
                r["scan_dir"], r["scan_dir_canonical"], r["scan_spelling_ok"], scan_url,
                r["scan_pages_url"], r["scan_repo_exists"], r["scan_repo_size_kb"],
                r["scan_repo_pushed_at"], r["scan_audit_note"], iso(r["public_link"]),
                r["in_pwgbib"], r["dict_citations"] if r["dict_citations"] is not None else "",
                r["dict_citations_key"],
                r["count_ratio"] if r["count_ratio"] is not None else "",
                "yes" if r["edition_qualified"] else "",
                r["ru_subset_refs"] if r["ru_subset_refs"] is not None else "",
                r["ru_subset_scan"] if r["ru_subset_scan"] is not None else "",
                r["ru_match_key"], r["resolver_repos"], r["scan_wired"],
                r["sheet_row_no"] or "",
            ])

    # ---- registry JSON ---------------------------------------------------------
    json_path = tracker_dir / "pwg_scan_index.json"
    payload = {
        "as_of": date.today().isoformat(),
        "source_sheet": SHEET_URL,
        "tracked_works": len(rows),
        "status_counts": {k: len(v) for k, v in sorted(by_status.items())},
        "citation_mass": {
            "tracked": tracked_mass,
            "indexed_done": indexed_mass,
            # The denominator these two may be divided by: the `ALL` of the same
            # ls-count snapshot they are summed from (H2874).
            "denominator": ls_all_tracker_era,
            "denominator_snapshot": "2024-09-11",
            # A different measurement of a different object, kept for scale only. It
            # counts cleaned citation strings, not work families -- never a denominator
            # for the two numbers above.
            "cleaned_string_occurrences_for_scale": crefs_total or None,
        },
        "works": [
            {k: (iso(v) if isinstance(v, date) else v) for k, v in r.items()}
            for r in rows
        ],
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=1) + "\n",
                         encoding="utf-8")

    # ---- e-text candidate queue (lane 3) --------------------------------------
    # Every done row is a candidate for text extraction. Rank by citation mass; the
    # RU-reuse column says whether the translated-subset pipeline would feel it too.
    queue_path = tracker_dir / "pwg_etext_candidate_queue.tsv"
    # Scan directories whose extraction question has already been settled elsewhere,
    # so the queue does not hand out work that is spoken for or already refuted.
    # Keyed by scan directory, not LS code, because a handoff takes a volume of a
    # book, not an abbreviation.
    ALREADY_QUEUED = {
        "ramayanabom": "assessed and REJECTED by H1705 (27-07-2026): the Bombay "
                       "uttarakāṇḍa has 111 sargas + 13 interpolated against the "
                       "corpus's 100, and the corpus file is 2,690 sa / 0 ru "
                       "critical-edition text — a Bombay concordance would have no "
                       "consumer. Do not re-derive.",
        "abch2": "OCR-from-scratch RULED OUT by the H1715 pilot (27-07-2026): local "
                 "tesseract 5 `san` scores 17.8% valid tokens against this work's own "
                 "committed e-text, while the Bayerische Staatsbibliothek already "
                 "publishes per-page hOCR for this edition at 43.8% — 2.5× better and "
                 "free. Re-scoped to ingest-and-correct; see reports/pwg_kosa_etext_pilot.md.",
        "amara_dlc": "same H1715 verdict — no text layer, but BSB publishes per-page "
                     "hOCR for this edition too (bsb10250868). Harvest, do not OCR.",
    }
    # Rank on the sheet's own count only, so the number that orders the queue is the
    # number the queue displays. Volumes of a multi-volume work carry no count of
    # their own (the sheet puts it on volume 1); they sort to the tail and are
    # marked, rather than being silently credited with a parent abbreviation's mass.
    cands = sorted(done_rows, key=lambda r: (-(r["citation_count"] or 0),
                                             -(r["total_pages"] or 0)))
    with queue_path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, delimiter="\t", lineterminator="\n")
        w.writerow(["rank", "ls_code", "title", "citation_count", "dict_citations",
                    "ru_subset_refs", "ru_subset_scan_hits", "total_pages", "scan_dir",
                    "scan_url", "scan_wired", "index_issue_url", "public_link",
                    "citations_per_page", "already_queued"])
        for i, r in enumerate(cands, 1):
            n = r["citation_count"] or r["dict_citations"] or 0
            pp = round(n / r["total_pages"], 1) if (n and r["total_pages"]) else ""
            w.writerow([
                i, r["ls_code"], r["title"], r["citation_count"] or "",
                r["dict_citations"] or "", r["ru_subset_refs"] or "",
                r["ru_subset_scan"] or "", r["total_pages"] or "", r["scan_dir"],
                f"{SCANS_ORG}/{r['scan_dir']}" if r["scan_dir"] else "",
                r["scan_wired"],
                (f"https://github.com/sanskrit-lexicon/{r['issue_repo']}/issues/{r['issue_no']}"
                 if r["issue_no"] else ""),
                iso(r["public_link"]), pp, ALREADY_QUEUED.get(r["scan_dir"], ""),
            ])

    # ---- site data -------------------------------------------------------------
    site_csv = site_data / "pwg_scan_index.csv"
    with site_csv.open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, lineterminator="\n")
        w.writerow(["ls_code", "title", "status", "vedic", "citations", "pages",
                    "volunteer", "finished", "public_link", "scan_dir"])
        for r in rows:
            w.writerow([r["ls_code"], r["title"], r["status"],
                        "yes" if r["vedic_marked"] else "no",
                        r["citation_count"] or "", r["total_pages"] or "",
                        first_handle(r["volunteer"]) if r["volunteer"] else "",
                        iso(r["finished"]), iso(r["public_link"]), r["scan_dir"]])

    summary = {
        "as_of": date.today().isoformat(),
        "source_sheet": SHEET_URL,
        "tracked_works": len(rows),
        "done": len(done_rows),
        "on_going": len(ongoing),
        "to_do": len(todo),
        "page_wise": len(pagewise),
        "not_required": len(nr),
        "volunteers": len(per_vol),
        "tracked_citation_mass": tracked_mass,
        "indexed_citation_mass": indexed_mass,
        "indexed_mass_pct_of_tracked": pct(indexed_mass, tracked_mass),
        # Dictionary-wide `<ls>` occurrence total, for scale ONLY. It is not a valid
        # denominator for indexed_citation_mass -- see report section 6.2.
        "dictionary_ls_occurrences_for_scale": crefs_total or None,
        # The denominators citation mass may legitimately be divided by (H2874): the
        # `ALL` of the ls-count snapshot the numerator came from. `citation_mass_*`
        # figures above are built from the 2024-09-11 column, so that is their partner.
        "citation_mass_denominator": ls_all_tracker_era,
        "citation_mass_denominator_snapshot": "2024-09-11",
        "citation_count_current_denominator": ls_all_current,
        "pages_indexed": pages(done_rows),
        "scans_published": sum(1 for r in done_rows if r["public_link"]),
        "scan_dirs_observed_wired": len({r["scan_dir"] for r in done_rows
                                         if r["scan_wired"] == "yes"}),
        "median_index_to_public_days": lag_median,
        "months": [{"month": m, "indexes_finished": fin_month.get(m, 0),
                    "scans_published": pub_month.get(m, 0)} for m in months],
        "per_volunteer": [
            {"handle": h, "works": v["rows"], "citation_mass": v["mass"], "pages": v["pages"]}
            for h, v in sorted(per_vol.items(), key=lambda kv: -kv[1]["mass"])
        ],
        "backlog": [
            {"ls_code": r["ls_code"], "title": r["title"], "citations": r["citation_count"],
             "pages": r["total_pages"], "vedic": r["vedic_marked"]} for r in todo
        ],
    }
    (site_data / "pwg_scan_index_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=1) + "\n", encoding="utf-8")

    # ---- report ----------------------------------------------------------------
    today = date.today().strftime("%d-%m-%Y")
    L = []
    A = L.append
    A("# PWG literary-source scan-index campaign — coverage by citation mass\n")
    A(f"_Created: 27-07-2026 · Last updated: {today}_\n")
    A(f"_Auto-generated by [`scripts/pwg_scan_index.py`]({REPO_BLOB}/scripts/pwg_scan_index.py) "
      "— Opus 5 (`claude-opus-5`), H1706. Source: MG's "
      f"[PWG literary-sources scan-index tracker]({SHEET_URL}) (snapshot committed under "
      f"[`data/pwg_scan_index_tracker/snapshot/`]({REPO_BLOB}/data/pwg_scan_index_tracker/snapshot)). "
      "Cross-checks read a local [sanskrit-lexicon/PWG](https://github.com/sanskrit-lexicon/PWG) "
      "checkout._\n")

    A("Between January 2025 and July 2026 a volunteer team page-indexed the printed "
      "editions that the Böhtlingk-Roth *Sanskrit-Wörterbuch* (PWG) cites, so that an "
      "`<ls>` citation in the dictionary can be resolved to the page image of the edition "
      "it cites. This report measures that campaign **by citation mass** — how much of the "
      "dictionary's citation apparatus the indexed works carry — because a 159-page "
      "*Kumārasaṃbhava* and a 2,420-page *Taittirīyabrāhmaṇa* are not equal work, and "
      "neither are equal payoff.\n")

    A("## 1 · Headline\n")
    A("| metric | value |")
    A("|---|--:|")
    A(f"| works tracked | {len(rows)} |")
    A(f"| indexed (`done`) | {len(done_rows)} |")
    A(f"| in progress (`on-going`) | {len(ongoing)} |")
    A(f"| unclaimed (`to-do`) | {len(todo)} |")
    A(f"| cited page-wise (no per-entry index needed) | {len(pagewise)} |")
    A(f"| not required (indirect / alternate name) | {len(nr)} |")
    A(f"| volunteers | {len(per_vol)} |")
    A(f"| pages indexed | {fmt(pages(done_rows))} |")
    A(f"| citation mass of tracked works | {fmt(tracked_mass)} |")
    A(f"| citation mass now indexed | {fmt(indexed_mass)} ({pct(indexed_mass, tracked_mass)}% of tracked) |")
    if lag_median is not None:
        A(f"| median index-posted → scan-public lag | {lag_median} days |")
    A("")
    A("**73.7% is coverage of the tracked set, not of the dictionary.** It answers \"how "
      "far has the campaign got through the works it took on\", and the denominator is "
      "the sheet's own citation-count column. It is deliberately **not** divided into the "
      "dictionary's total citation count: the two numbers are not commensurable, and §6 "
      "shows why. The long tail of thousands of once-cited works was never in scope for "
      "indexing anyway.\n")

    A("## 2 · Status distribution\n")
    A("The sheet's own status vocabulary, normalized. `page-wise` and the two `NR` values "
      "are **not** backlog: they are rulings that a per-entry index is the wrong "
      "instrument for that work.\n")
    A("| status | meaning | works | citation mass | pages |")
    A("|---|---|--:|--:|--:|")
    order = ["done", "on-going", "to-do", "page-wise", "nr-indirect", "nr-alt-name"]
    for st in order + [s for s in sorted(by_status) if s not in order]:
        rs = by_status.get(st)
        if not rs:
            continue
        gloss = rs[0]["status_gloss"] or "—"
        A(f"| `{st}` | {gloss} | {len(rs)} | {fmt(mass(rs))} | {fmt(pages(rs))} |")
    A("")

    A("## 3 · What remains — the ranked backlog\n")
    A(f"**{len(todo)} unclaimed works** carrying **{fmt(mass(todo))}** citations "
      f"({pct(mass(todo), tracked_mass)}% of tracked mass) across "
      f"{fmt(pages(todo))} pages. Ranked by citation payoff:\n")
    A("| # | LS code | work | citations | pages | citations/page |")
    A("|--:|---|---|--:|--:|--:|")
    for i, r in enumerate(todo, 1):
        n, p = r["citation_count"] or 0, r["total_pages"] or 0
        A(f"| {i} | `{r['ls_code']}` | {r['title']} | {fmt(n)} | {fmt(p)} | "
          f"{round(n/p, 1) if p else '—'} |")
    A("")
    if ongoing:
        A(f"In progress: " + ", ".join(
            f"`{r['ls_code']}` ({fmt(r['citation_count'] or 0)} citations, "
            f"{r['volunteer'] or 'unassigned'})" for r in ongoing) + ".\n")

    vedic_todo = [r for r in todo if r["vedic_marked"]]
    if vedic_todo:
        A(f"**The backlog is Vedic.** {len(vedic_todo)} of the {len(todo)} unclaimed works "
          "carry the sheet's ★ mark, and the marked set is exactly the Vedic "
          "saṃhitā / brāhmaṇa / upaniṣad / śrauta- and gṛhya-sūtra / prātiśākhya group — "
          "long texts with dense citation and awkward reference schemes. The kāvya and "
          "kośa material, which indexes fast, is essentially finished.\n")

    A("## 4 · Per-volunteer throughput\n")
    A("Attribution exactly as the sheet records it — the `Reserved/Indexed by` column, "
      "first handle where a row carries two. Rows include the `NR/alt. name` and "
      "`on-going` work a volunteer was assigned, not only completed indexes.\n")
    A("| volunteer | works | citation mass | pages |")
    A("|---|--:|--:|--:|")
    for h, v in sorted(per_vol.items(), key=lambda kv: -kv[1]["mass"]):
        A(f"| [{h}](https://github.com/{h.lstrip('@')}) | {v['rows']} | {fmt(v['mass'])} | {fmt(v['pages'])} |")
    A("")

    A("## 5 · Velocity\n")
    A("Two curves: when an index was **finished** by its volunteer, and when the scan "
      "directory went **public**. They are deliberately separate — the second is a "
      "different pipeline (upload, review, publish) with its own queue.\n")
    A("| month | indexes finished | scans published |")
    A("|---|--:|--:|")
    for m in months:
        A(f"| {m} | {fin_month.get(m, 0) or '—'} | {pub_month.get(m, 0) or '—'} |")
    A("")
    if lags:
        A(f"Index-posted → public-link lag over {len(lags)} works: median **{lag_median} days**, "
          f"range {lags[0]}–{lags[-1]} days.\n")

    A("## 6 · Cross-validation against the dictionary's own data\n")
    A("The tracker is a human work-log; the PWG repository holds independently derived "
      "citation data. Two checks, with very different outcomes.\n")

    tiers = Counter(r["pwgbib_tier"] for r in rows)
    A("### 6.1 LS codes — clean\n")
    A(f"Every tracked LS entry code was resolved against the PWG abbreviation bibliography "
      f"[`pwgbib_input.txt`]({PWG_BLOB}/pwg_ls1/pwgauth/pwgbib_input.txt) "
      f"({len(pwgbib)} abbreviations), cheapest test first.\n")
    A("| match | rows | what it means |")
    A("|---|--:|---|")
    A(f"| verbatim | {tiers.get('exact', 0)} | sheet and bibliography spell it identically |")
    A(f"| case only | {tiers.get('case', 0)} | e.g. the sheet's `MBh.` vs the bibliography's `MBH.` |")
    A(f"| after dropping the sheet's edition/volume qualifier | "
      f"{tiers.get('qualified', 0) + tiers.get('qualified-case', 0)} | "
      "`Spr. vol.1 (1st ed.)` → `Spr.`, `AK. Deslongchamps ed.` → `AK.` |")
    A(f"| unresolved | {tiers.get('none', 0)} | |")
    A("")
    loose = sorted({r["ls_code"] for r in rows if r["pwgbib_tier"] == "none"})
    if loose:
        A(f"The {len(loose)} unresolved: "
          + "; ".join(f"`{c}`" for c in loose)
          + " — recorded as a loose end, not explained away. Each names a real work the "
            "campaign indexed; what is unresolved is only whether the bibliography lists "
            "that exact abbreviation.\n")
    else:
        A("**Nothing unresolved.** Every tracked row names a work the dictionary's own "
          "bibliography lists.\n")

    A("### 6.2 Citation counts — provenance resolved, and still not commensurable\n")
    A("The sheet's `Citation count` column is the per-abbreviation total of "
      f"[`pwgissues/issue74/lsextract_all.txt`]({PWG_BLOB}/pwgissues/issue74/lsextract_all.txt), "
      f"the PWG count table dated **2024-09-11** with `ALL = {fmt(ls_all_tracker_era)}`: every "
      "`<ls>` element in the dictionary attributed to the longest bibliography abbreviation "
      "that prefixes it. 66 of the 67 valued rows match it digit for digit. That table is "
      "committed here as the input of record, the column is regenerable by "
      f"[`scripts/pwg_ls_counts.py`]({REPO_BLOB}/scripts/pwg_ls_counts.py), and the full "
      "argument plus the per-row dataset is in "
      f"[`reports/pwg_citation_count_provenance.md`]({REPO_BLOB}/reports/pwg_citation_count_provenance.md) "
      "(H2874).\n")
    A("So the counts **do** have a denominator now — their own snapshot's `ALL`, carried in "
      "the summary JSON as `citation_mass_denominator`. What they still are not is "
      "commensurable with the cleaned-string extraction below, and rows of one "
      "bibliography family still share a single total, so a naive row sum double-counts. "
      "The registry therefore exposes `citation_count_safe` (the source table's own number, "
      "blank where provenance is unresolved) alongside the sheet's `citation_count`, and "
      "every figure in this report is built from the safe field.\n")
    have_both = [r for r in rows if r["count_ratio"]]
    A("The sheet's `Citation count` column and the full-dictionary `<ls>` extraction "
      f"([`sortedcrefs.txt`]({PWG_BLOB}/pwg_ls/pwg_dhaval/abbrvwork/abbrvoutput/sortedcrefs.txt)) "
      "count **different objects**, so they cannot be reconciled row by row and no attempt "
      "is made to. The extraction keys on a *cleaned citation string* — canto and śloka "
      "numbers stripped, everything else kept — so one book scatters across many keys "
      "(`MED.`, `MED. k.`, `MED. kh.`, `MED. im ŚKDR.` …). The sheet keys on a *book*.\n")
    if have_both:
        ratios = sorted(r["count_ratio"] for r in have_both)
        A(f"Over the {len(have_both)} rows where both numbers exist, the sheet/extraction "
          f"ratio ranges from **{ratios[0]}× to {ratios[-1]}×** with median "
          f"**{ratios[len(ratios)//2]}×** — a spread far too wide for the two to be "
          "measuring one quantity with noise.\n")
        A("| LS code | sheet | bare-string occurrences in `sortedcrefs` | ratio |")
        A("|---|--:|--:|--:|")
        for r in sorted(have_both, key=lambda r: -r["count_ratio"])[:8]:
            A(f"| `{r['ls_code']}` | {fmt(r['citation_count'])} | {fmt(r['dict_citations'])} | "
              f"{r['count_ratio']}× |")
        for r in sorted(have_both, key=lambda r: r["count_ratio"])[:3]:
            A(f"| `{r['ls_code']}` | {fmt(r['citation_count'])} | {fmt(r['dict_citations'])} | "
              f"{r['count_ratio']}× |")
        A("")
    A(f"For scale, the extraction totals **{fmt(crefs_total)}** `<ls>` occurrences over "
      f"{fmt(len(crefs))} distinct cleaned strings across the whole dictionary.\n")
    A("> **Closed, with one residue.** H1706 recorded this column's provenance as open and "
      "barred it from any denominator; H2874 traced it to the 2024-09-11 count table above "
      "and replaced the ban with a contract "
      f"(`python scripts/pwg_citation_count_provenance.py --check`). The residue is "
      "`NAIGH.`, whose sheet cell reads 1,477 against 1,417 in the source table and matches "
      "no committed snapshot of the count — carried as a near-miss, never silently "
      "corrected. What remains genuinely unrecorded is *who* ran the extraction and when "
      "they pasted it in; the number no longer depends on that testimony.\n")
    # De-duplicate by the matched abbreviation: several volume rows share one parent
    # count, and listing that count three times would read as three works.
    ru_by_key = {}
    for r in rows:
        if r["ru_subset_refs"] and r["ru_match_key"] not in ru_by_key:
            ru_by_key[r["ru_match_key"]] = r["ru_subset_refs"]
    if ru_by_key:
        top = sorted(ru_by_key.items(), key=lambda kv: -kv[1])[:5]
        A("A third count exists and is smaller by construction: the translated-article "
          f"subset ([`citation_sources.json`]({PWG_BLOB}/pwg_ls/pwg_ru_coverage/citation_sources.json)) "
          f"resolves {len(ru_by_key)} of these abbreviations, the heaviest being "
          + ", ".join(f"`{k}` {fmt(v)}" for k, v in top)
          + ". It measures ~51 translated roots, not the dictionary, and is reported in the "
            "registry for reuse-value ranking only. Where a row's own code is not in that "
            "measurement the registry records the parent abbreviation it matched, in "
            "`ru_match_key` — volume rows sharing a parent must not be summed.\n")

    A("## 7 · The frontier — published, wired, and worth extracting\n")
    published = [r for r in done_rows if r["public_link"]]
    A("An index pays off only after three gates, and they are not the same count: the "
      "index is **finished**, the scan directory is **public**, and the citation "
      "resolver is **wired** to emit a link to it.\n")
    A("| gate | count |")
    A("|---|--:|")
    A(f"| index finished (`done`) | {len(done_rows)} works |")
    A(f"| scan directory public (dated `Public Link`) | {len(published)} works |")
    if scan_audit:
        audit_rows = {v["scan_dir"]: v for v in scan_audit.values()}
        vcount = Counter(v["resolver_wired"] for v in audit_rows.values())
        A(f"| scan repository exists and serves GitHub Pages | "
          f"{sum(1 for v in audit_rows.values() if v['repo_exists'] == 'true')} / "
          f"{len(audit_rows)} directories |")
        A(f"| resolver fully wired | {vcount.get('yes', 0)} directories |")
        A(f"| resolver partially wired (some citation arities return nothing) | "
          f"{vcount.get('partial', 0)} |")
        A(f"| resolver mis-keyed (target exists, unreachable from the tracked code) | "
          f"{vcount.get('mis-keyed', 0)} |")
        A(f"| resolver not wired at all | {vcount.get('no', 0)} |")
        A("")
        A("Verified by a dated static audit of the resolver source and the live GitHub org "
          "— see [`scan_target_audit.tsv`]"
          f"({REPO_BLOB}/data/pwg_scan_index_tracker/scan_target_audit.tsv). "
          f"All {len(audit_rows)} directories are real public repositories under "
          "[sanskrit-lexicon-scans](https://github.com/sanskrit-lexicon-scans), each serving "
          "a page-lookup app at `https://sanskrit-lexicon-scans.github.io/<dir>/`; together "
          "they hold about 11.2 GB of page images.\n")
        problems = [v for v in sorted(audit_rows.values(), key=lambda v: v["scan_dir"])
                    if v["resolver_wired"] != "yes" or v["spelling_matches"] != "yes"]
        if problems:
            A("### Defects this audit found\n")
            A("Each is a concrete, checkable claim about a specific line of the resolver or a "
              "specific tracker cell — not a heuristic guess.\n")
            for v in problems:
                A(f"- **`{v['scan_dir']}`** — {v['resolver_wired']}"
                  + ("; tracker spells it "
                     f"`{v['tracker_spelling']}`, which GitHub Pages will not serve"
                     if v["spelling_matches"] != "yes" else "")
                  + f". {v['note']}")
            A("")
            A("> **The `rvps` case is a wrong link, not a missing one.** A Ṛgveda-Prātiśākhya "
              "citation does not fail to resolve — it resolves to an Ṛgveda *hymn* anchor, a "
              "different text. A silent wrong answer is worse than a visible gap, so this one "
              "is filed as an integrity defect rather than left in a backlog.\n")
    else:
        wired = [r for r in done_rows if r["scan_wired"] == "observed"]
        A(f"| scan directory observed in the citation resolver | "
          f"{len({r['scan_dir'] for r in wired})} directories |")
        A("")
        A("_No static audit file present; the wiring column falls back to what the "
          "translated-subset measurement happened to observe, which is weak evidence._\n")
    A("### E-text candidates\n")
    A("Every finished index is a candidate for full-text extraction: the per-page index is "
      "exactly the segmentation anchor a page-image OCR pass needs. Ranked by citation "
      "payoff — the full list is "
      f"[`pwg_etext_candidate_queue.tsv`]({REPO_BLOB}/data/pwg_scan_index_tracker/pwg_etext_candidate_queue.tsv).\n")
    A("| # | LS code | work | citations | pages | scan directory | note |")
    A("|--:|---|---|--:|--:|---|---|")
    for i, r in enumerate(cands[:15], 1):
        notes_q = []
        if r["scan_dir_canonical"] in ALREADY_QUEUED:
            notes_q.append(ALREADY_QUEUED[r["scan_dir_canonical"]])
        if r["scan_wired"] in ("no", "partial", "mis-keyed"):
            notes_q.append(f"resolver {r['scan_wired']}")
        A(f"| {i} | `{r['ls_code']}` | {r['title']} | "
          f"{fmt(r['citation_count']) if r['citation_count'] else '—'} | "
          f"{fmt(r['total_pages']) if r['total_pages'] else '—'} | "
          + (f"[`{r['scan_dir_canonical']}`]({SCANS_ORG}/{r['scan_dir_canonical']})"
             if r["scan_dir_canonical"] else "—")
          + f" | {'; '.join(notes_q) or '—'} |")
    A("")
    A("Settled elsewhere, so **not** available to a new handoff:\n")
    for k, v in ALREADY_QUEUED.items():
        A(f"- `{k}` — {v}")
    A("")

    A("## 8 · Data-quality notes on the sheet itself\n")
    notes = []
    dup_books = [b for b, c in Counter(r["book_no"] for r in rows if r["book_no"]).items() if c > 1]
    if dup_books:
        dups = ", ".join(sorted(dup_books))
        notes.append(f"**Book-number collisions.** `{dups}` — the same `Book No.` is used by "
                     "more than one row, so it is not a key. The registry keys on the LS code.")
    sheet_nos = [r["sheet_row_no"] for r in rows if r["sheet_row_no"]]
    if sheet_nos:
        gaps = sorted(set(range(min(sheet_nos), max(sheet_nos) + 1)) - set(sheet_nos))
        if gaps:
            notes.append(f"**Serial-number gaps.** The sheet's own `S. No.` column skips "
                         f"{', '.join(str(g) for g in gaps)} — rows were deleted after numbering.")
    if any(r["scan_dir_raw"].startswith("/") for r in rows):
        bad = [r["ls_code"] for r in rows if r["scan_dir_raw"].startswith("/")]
        notes.append(f"**Leading-slash typo** in the scan-directory cell for "
                     f"{', '.join('`'+c+'`' for c in bad)}; normalized in the registry.")
    counts_csv = parse_simple_csv(snapshot_dir / "counts.csv")
    if counts_csv:
        summary_months = [c[1] for c in counts_csv if len(c) > 2 and c[1].startswith("(20")]
        if summary_months and months:
            last_summary = summary_months[-1]
            notes.append(f"**The sheet's own summary tabs are stale.** Its `Counts` tab stops at "
                         f"{last_summary}, while the `Works` tab records work through "
                         f"{months[-1]}. Every number in this report is recomputed from `Works`; "
                         "the summary tabs are snapshotted but not used.")
    finished_csv = parse_simple_csv(snapshot_dir / "works_finished.csv")
    if finished_csv:
        wf_titles = {c[1].strip() for c in finished_csv if len(c) > 1 and c[1].strip()}
        works_titles = {r["title"] for r in rows}
        orphan = sorted(t for t in wf_titles
                        if t and t != "Complete" and not any(t.lower() in w.lower() or w.lower() in t.lower()
                                                             for w in works_titles))
        if orphan:
            notes.append("**`Works finished` lists works absent from `Works`** — "
                         + ", ".join(f"*{t}*" for t in orphan)
                         + ". Either a plan that never became a tracked row, or a row deleted "
                           "from `Works`; the registry follows `Works`.")
    notes.append("**One redaction in the committed snapshot.** The sheet's `Team` tab maps "
                 "each volunteer's real personal name to their GitHub handle. The handles "
                 "are public — they are on every coordinating issue — but that linkage is "
                 "not, and this repository is. The name column is redacted at fetch time; "
                 "the handle column, and therefore every credit and count in this report, "
                 "is untouched.")
    no_count = [r for r in rows if r["citation_count"] is None]
    if no_count:
        notes.append(f"**{len(no_count)} rows carry no citation count** (the sheet writes "
                     "`- - -` where a multi-volume work's count sits on its first volume). "
                     "Their mass is therefore attributed to the parent row, and per-volume "
                     "mass figures in this report are floors, not exact.")
    for n in notes:
        A(f"- {n}")
    A("")

    A("## 9 · Files\n")
    A(f"- registry — [`data/pwg_scan_index_tracker/pwg_scan_index.tsv`]({REPO_BLOB}/data/pwg_scan_index_tracker/pwg_scan_index.tsv) "
      f"· [`.json`]({REPO_BLOB}/data/pwg_scan_index_tracker/pwg_scan_index.json)")
    A(f"- e-text candidate queue — [`pwg_etext_candidate_queue.tsv`]({REPO_BLOB}/data/pwg_scan_index_tracker/pwg_etext_candidate_queue.tsv)")
    A(f"- sheet snapshot (all four tabs, verbatim) — [`snapshot/`]({REPO_BLOB}/data/pwg_scan_index_tracker/snapshot)")
    A(f"- campaign history — [`docs/PWG_SCAN_INDEX_CAMPAIGN_HISTORY_2025_2026.md`]({REPO_BLOB}/docs/PWG_SCAN_INDEX_CAMPAIGN_HISTORY_2025_2026.md)")
    A(f"- complementary metric (how many citations *link out*, translated subset) — "
      f"[`reports/pwg_citation_coverage.md`]({REPO_BLOB}/reports/pwg_citation_coverage.md)")
    A("")
    A("_Dr. Mārcis Gasūns_\n")

    report_path = reports_dir / "pwg_scan_index.md"
    report_path.write_text("\n".join(L), encoding="utf-8")

    print(f"Wrote {reg_path}")
    print(f"Wrote {json_path}")
    print(f"Wrote {queue_path}")
    print(f"Wrote {site_csv}")
    print(f"Wrote {site_data / 'pwg_scan_index_summary.json'}")
    print(f"Wrote {report_path}")
    print(f"works={len(rows)} done={len(done_rows)} todo={len(todo)} "
          f"mass_indexed={indexed_mass:,}/{tracked_mass:,} volunteers={len(per_vol)}")
    if crefs_total:
        print(f"dictionary <ls> occurrences (scale only, NOT a denominator): {crefs_total:,}")
    else:
        print(f"WARNING: no sortedcrefs at {crefs_path}; dictionary-wide share not computed",
              file=sys.stderr)
    if not pwgbib:
        print(f"WARNING: no pwgbib at {pwgbib_path}; LS-code validation skipped", file=sys.stderr)
    if not ru_counts:
        print(f"WARNING: no translated-subset counts at {ru_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
